from app.models import querySala,queryGalen
from openpyxl import Workbook
from openpyxl.styles import colors,PatternFill,Font

def generarReporteSala(fechai,fechas):
	try:
		wb=Workbook()
		ws=wb.active	
		obj_sala=querySala()
		obj_galen=queryGalen()
		rows=obj_sala.consultarTablaDate('Sala','Fech_Ingre',fechai,fechas)
		nro=2
		columna=['Historia Clinica','Documento','Nombres','ApellidoPaterno','ApellidoMaterno','Departamento','Provincia','Distrito',
		'Interv Quirurgica','UPSS','Complejidad','DX1','DX2','Procedimiento','DX POS 1','DX POS 2','Interv','Cirujano','Anestesiologo',
		'Instrumentista I','InstrumentistaII','Fecha Ingreso','Hora Ingreso','Fecha Salida','Hora Salida','Tiempo Uso','Destino','Finaciador',
		'Responsable']
		letra_columna=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC']
		ws.merge_cells('A1:AC1')
		ws['A1']=f"REPORTE GENERAL DE PACIENTES REGISTRADOS EN SALA DESDE {fechai} HASTA {fechas}"
		for i in range(len(columna)):
			ws[letra_columna[i]+str(nro)]=columna[i]
			ws[letra_columna[i]+str(nro)].font = Font(color = 'FF0000',bold=True, size=12)
			ws[letra_columna[i]+str(nro)].fill=PatternFill('solid', start_color="38e3ff")
			ws.column_dimensions[letra_columna[i]].width = 30
		#llenando datos
		nro_start=3
		for val in rows:
			historia=val.HC
			ws[letra_columna[0]+str(nro_start)]=historia

			rows_pac=obj_galen.consultarDatosPaciente(historia)

			if rows_pac:
				ws[letra_columna[1]+str(nro_start)]=rows_pac[0].NroDocumento
				ws[letra_columna[2]+str(nro_start)]=rows_pac[0].PrimerNombre
				ws[letra_columna[3]+str(nro_start)]=rows_pac[0].ApellidoPaterno
				ws[letra_columna[4]+str(nro_start)]=rows_pac[0].ApellidoMaterno
				ws[letra_columna[5]+str(nro_start)]=rows_pac[0].DEPARTAMENTO
				ws[letra_columna[6]+str(nro_start)]=rows_pac[0].PROVINCIA
				ws[letra_columna[7]+str(nro_start)]=rows_pac[0].DISTRITO


			ws[letra_columna[8]+str(nro_start)]=val.Inter_Q
			ws[letra_columna[9]+str(nro_start)]=val.UPPS
			ws[letra_columna[10]+str(nro_start)]=val.COMPLEJIDAD
			ws[letra_columna[11]+str(nro_start)]=val.DX1
			ws[letra_columna[12]+str(nro_start)]=val.DX2
			ws[letra_columna[13]+str(nro_start)]=val.INTQ
			ws[letra_columna[14]+str(nro_start)]=val.DX1POST
			ws[letra_columna[15]+str(nro_start)]=val.DX2POST
			ws[letra_columna[16]+str(nro_start)]=val.R_Interv
			ws[letra_columna[17]+str(nro_start)]=val.Cirujano
			ws[letra_columna[18]+str(nro_start)]=val.Anestesiologo
			ws[letra_columna[19]+str(nro_start)]=val.InstrumentistaI
			ws[letra_columna[20]+str(nro_start)]=val.InstrumentistaII
			ws[letra_columna[21]+str(nro_start)]=val.Fech_Ingre
			ws[letra_columna[22]+str(nro_start)]=val.Hora_Ingre
			ws[letra_columna[23]+str(nro_start)]=val.Fech_Sali
			ws[letra_columna[24]+str(nro_start)]=val.Hora_Sali
			ws[letra_columna[25]+str(nro_start)]=val.Tiempo_Uso
			ws[letra_columna[26]+str(nro_start)]=val.Servicio_Deriva
			ws[letra_columna[27]+str(nro_start)]=val.Financiamiento
			ws[letra_columna[28]+str(nro_start)]=val.Responsable
			nro_start=nro_start+1			
		return wb

	except Exception as e:
		print(e)

def generarReporteUrpa(fechai,fechas):
	try:
		wb=Workbook()
		ws=wb.active	
		obj_sala=querySala()
		obj_galen=queryGalen()
		rows=obj_sala.consultarTablaDate('URPA','FECHAI',fechai,fechas)
		nro=2
		columna=['Historia Clinica','Documento','Nombres','ApellidoPaterno','ApellidoMaterno','Departamento','Provincia','Distrito',
		'Fecha Ingreso','Hora Ingreso','Fecha Salidad','Hora Salida','Tiempo Permanencia','Tiempo Espera','Destino','Responsable','Anestesia']
		letra_columna=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
		ws.merge_cells('A1:Q1')
		ws['A1']=f"REPORTE GENERAL DE PACIENTES REGISTRADOS EN URPA DESDE {fechai} HASTA {fechas}"
		for i in range(len(columna)):
			ws[letra_columna[i]+str(nro)]=columna[i]
			ws[letra_columna[i]+str(nro)].font = Font(color = 'FF0000',bold=True, size=12)
			ws[letra_columna[i]+str(nro)].fill=PatternFill('solid', start_color="38e3ff")
			ws.column_dimensions[letra_columna[i]].width = 30
		#llenando datos
		nro_start=3
		for val in rows:
			idsala=val.Id_Sala
			rows_sala=obj_sala.consultarTablaCondition('Sala','Id_Sala',idsala)
			historia=rows_sala[0].HC
			ws[letra_columna[0]+str(nro_start)]=historia

			rows_pac=obj_galen.consultarDatosPaciente(historia)

			if rows_pac:
				ws[letra_columna[1]+str(nro_start)]=rows_pac[0].NroDocumento
				ws[letra_columna[2]+str(nro_start)]=rows_pac[0].PrimerNombre
				ws[letra_columna[3]+str(nro_start)]=rows_pac[0].ApellidoPaterno
				ws[letra_columna[4]+str(nro_start)]=rows_pac[0].ApellidoMaterno
				ws[letra_columna[5]+str(nro_start)]=rows_pac[0].DEPARTAMENTO
				ws[letra_columna[6]+str(nro_start)]=rows_pac[0].PROVINCIA
				ws[letra_columna[7]+str(nro_start)]=rows_pac[0].DISTRITO

			ws[letra_columna[8]+str(nro_start)]=val.FECHAI
			ws[letra_columna[9]+str(nro_start)]=val.HORAI
			ws[letra_columna[10]+str(nro_start)]=val.FECHAS
			ws[letra_columna[11]+str(nro_start)]=val.HORAS
			ws[letra_columna[12]+str(nro_start)]=val.T_PERMANENCIA
			ws[letra_columna[13]+str(nro_start)]=val.T_ESPERA
			ws[letra_columna[14]+str(nro_start)]=val.DERIVAR
			ws[letra_columna[15]+str(nro_start)]=val.RESPONSABLE
			rows_anestesia=obj_sala.consultarTablaCondition('TANESTECIA','Id_TAnestecia',val.Id_TAnestecia)
			ws[letra_columna[16]+str(nro_start)]=rows_anestesia[0].tipoA
			
			nro_start=nro_start+1			
		return wb

	except Exception as e:
		print(e)

	